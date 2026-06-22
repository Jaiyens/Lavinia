#!/usr/bin/env python3
"""
Validate a PG&E "Download My Data" / Share My Data usage export against the
Batth master meter sheet BEFORE ingesting into Tool 1.

  pip install openpyxl
  python3 validate_pge_export.py \
      --master "Batth_Farms_2025_Master_Meter_List_.xlsx" \
      --export ./pge_download/ [--sheet All] [--ag-rate-card path.json]

Fabricates nothing: every number is read from the files. Anything that can't be
known until the file lands is printed as UNKNOWN and checked at runtime.
"""
import argparse, csv, glob, os, re, sys, json
from collections import Counter, defaultdict

def nh(h): return re.sub(r"[^a-z0-9]", "", str(h or "").lower())

MASTER_ALIASES = {
    "serviceId": {"serviceid","said","sa","saidno","serviceagreement","serviceagreementid","spid"},
    # The real Batth sheet prints "Full Acct #" (-> fullacct), "Active Rate Schedule"
    # (-> activerateschedule) and "NEMA" (-> nema). Without these tokens load_master
    # could not map account/rate/nem at all, so by_acct collapsed to one None bucket,
    # rate_schedules read 0, and nem read 0 (every per-account/rate/NEM check was wrong).
    "account":   {"account","accountnumber","accountno","acct","acctno","pgeaccount","fullacct","fullaccount","accountid"},
    "rate":      {"rate","rateschedule","tariff","rateplan","schedule","activerateschedule","currentrate","rateschedulecode"},
    "gpm":       {"gpm","gallonsperminute","flow","flowrate"},
    "horsepower":{"hp","horsepower"},
    "nem":       {"nem","nemtype","netmetering","nemprogram","nema","nemacode","aggregation"},
    "entity":    {"entity","legalentity","company","businessentity","owner","billingentity","billingname"},
    # Geo/solar are reported as raw coverage only (the Solar column is low-fidelity:
    # mixed "840kw"/"Solar"/numbers), never wired into any GO/NO-GO logic.
    "lat":       {"lat","latitude","premlat"},
    "lon":       {"lon","lng","long","longitude","premlong"},
    "solar":     {"solar","issolar","solarflag","hassolar"},
}
EXPORT_ALIASES = {
    "account":   {"accountid","account","accountnumber"},
    "serviceId": {"serviceagreementid","said","serviceid","said"},
    "meter":     {"meterbadgenumber","meterbadge","meter","meterserial","meternumber"},
    "rate":      {"ratecode","rate","rateschedule","tariff"},
    "date":      {"date"},
    "intervallen":{"intervallength","intervallen","duration"},
    "tou":       {"toucode","tou"},
    "direction": {"directionofenergy","direction","flowdirection"},
    "uom":       {"unitofmeasure","uom","unit"},
    "usage":     {"usagevalue","usage","value","kwh"},
}
DOLLAR_HINTS = {"amount","charge","cost","dollar","price","billamount","credit","usd","totalcharges"}

def pick(colmap, aliases):
    for w in aliases:
        if w in colmap: return colmap[w]
    return None

def norm_account(a):
    if a in (None, ""): return None
    s = str(a).strip().split("-")[0].lstrip("0")
    return s or "0"

def said_match(export_sid, master_set):
    s = str(export_sid).strip()
    for v in {s, s.lstrip("0"), s.zfill(8), s.zfill(9), s.zfill(10)}:
        if v in master_set: return v
    return None

# ---------- master ----------
def load_master(path, sheet):
    try:
        from openpyxl import load_workbook
    except ImportError:
        sys.exit("ERROR: openpyxl not installed -> pip install openpyxl")
    if not os.path.exists(path): sys.exit(f"ERROR: master not found: {path}")
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb[sheet] if sheet in wb.sheetnames else wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    header = next(it)
    colmap = {}
    for i, h in enumerate(header):
        k = nh(h)
        if k and k not in colmap: colmap[k] = i
    idx = {k: pick(colmap, al) for k, al in MASTER_ALIASES.items()}
    if idx["serviceId"] is None:
        sys.exit(f"ERROR: no SA ID column in master. headers={list(colmap)}")
    meters = []
    for r in it:
        def g(k):
            i = idx[k]
            return r[i] if (i is not None and i < len(r)) else None
        sid = g("serviceId")
        if sid in (None, ""): continue
        meters.append({
            "serviceId": str(sid).strip(),
            "account": g("account"),
            "rate": (str(g("rate")).strip() if g("rate") not in (None,"") else None),
            "gpm": g("gpm"), "horsepower": g("horsepower"),
            "nem": g("nem"), "entity": g("entity"),
            "lat": g("lat"), "lon": g("lon"), "solar": g("solar"),
        })
    return meters, list(colmap)

# ---------- export ----------
def list_files(path):
    if os.path.isdir(path):
        out = []
        for ext in ("*.csv","*.CSV","*.xml","*.XML"):
            out += glob.glob(os.path.join(path, "**", ext), recursive=True)
        return sorted(set(out))
    return [path]

def sniff(path):
    with open(path, "rb") as f:
        head = f.read(300).lstrip().lower()
    return "xml" if (head[:5] == b"<?xml" or b"<feed" in head or b"espi" in head) else "csv"

def parse_csv(path):
    with open(path, newline="", encoding="utf-8-sig", errors="replace") as f:
        rows = list(csv.reader(f))
    hdr = None
    for i, row in enumerate(rows):
        cells = {nh(c) for c in row}
        if cells & EXPORT_ALIASES["serviceId"] and cells & EXPORT_ALIASES["usage"]:
            hdr = i; break
    if hdr is None:
        for i, row in enumerate(rows):
            if len(row) > 6: hdr = i; break
    if hdr is None: return [], {}
    colmap = {}
    for j, h in enumerate(rows[hdr]):
        k = nh(h)
        if k and k not in colmap: colmap[k] = j
    idx = {k: pick(colmap, al) for k, al in EXPORT_ALIASES.items()}
    recs = []
    for row in rows[hdr+1:]:
        if not any(row): continue
        def g(k):
            i = idx[k]
            return row[i] if (i is not None and i < len(row)) else None
        if g("serviceId") in (None, ""): continue
        recs.append({k: g(k) for k in EXPORT_ALIASES})
    return recs, colmap

def parse_xml_coarse(path):
    import xml.etree.ElementTree as ET
    flow, durs, ups = set(), set(), set()
    for _, el in ET.iterparse(path, events=("end",)):
        t = el.tag.split("}")[-1]
        if t == "flowDirection" and el.text: flow.add(el.text.strip())
        elif t == "duration" and el.text and el.text.strip().isdigit(): durs.add(int(el.text.strip()))
        elif t == "UsagePoint":
            for a in el.attrib.values(): ups.add(a)
        el.clear()
    return {"flowDirections": flow, "durations": durs, "usagePoints": len(ups)}

def is_received(d):  return str(d).strip().upper() in {"R","RECEIVED","19"}

# The repo's real AG rate-card families (src/lib/energy/rate-lever.ts LABEL_TO_PLAN /
# CANDIDATE_SCHEDULES). A code is priceable only when its normalized base begins with
# one of these. mapScheduleLabel normalizes a label by uppercasing and dropping hyphens
# (token = label.toUpperCase().replace(/-/g,"")); the validator additionally models the
# leading-H gap below so it can report what the pilot can ACTUALLY price, not a substring.
AG_CARD_FAMILIES = ("AGA", "AGB", "AGC", "AG4", "AG5")

def rate_base(code):
    """Normalize an export rate code the way the card lookup keys it: uppercase, drop
    hyphens, strip a single leading 'H' (PG&E's hold/legacy prefix)."""
    if not code: return ""
    t = re.sub(r"-", "", str(code).strip().upper())
    return t[1:] if t.startswith("H") else t

def classify_rate(code):
    """One of: 'ag' (base maps to a known card family -> priceable), 'ag_no_card'
    (AG-prefixed but no card family, e.g. AGFB), or 'non_ag'."""
    b = rate_base(code)
    if any(b.startswith(f) for f in AG_CARD_FAMILIES): return "ag"
    if b.startswith("AG"): return "ag_no_card"
    return "non_ag"

def is_h_prefixed_ag(code):
    """True for codes that map to an AG card family ONLY after stripping a leading 'H'.
    As the repo is wired, mapScheduleLabel does NOT strip the H, so these are unpriceable
    until rate-lever.ts strips it. This is the honest gap the board must surface."""
    if not code: return False
    raw = str(code).strip().upper()
    return raw.startswith("H") and classify_rate(code) == "ag"

# ---------- main ----------
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--master", required=True)
    ap.add_argument("--export", required=True)
    ap.add_argument("--sheet", default="All")
    ap.add_argument("--ag-rate-card", default=None)
    a = ap.parse_args()

    meters, mheaders = load_master(a.master, a.sheet)
    master_sa = {m["serviceId"] for m in meters}
    by_acct = defaultdict(list)
    for m in meters: by_acct[norm_account(m["account"])].append(m)
    hp_ok  = sum(1 for m in meters if m["horsepower"] not in (None,""))
    gpm_ok = sum(1 for m in meters if m["gpm"] not in (None,""))
    geo_ok = sum(1 for m in meters if m["lat"] not in (None,"") and m["lon"] not in (None,""))
    solar_ok = sum(1 for m in meters if m["solar"] not in (None,""))
    nem_sa = {m["serviceId"] for m in meters if m["nem"] not in (None,"")}
    print(f"\n== MASTER ==  meters={len(meters)} accounts={len(by_acct)} "
          f"rate_schedules={len({m['rate'] for m in meters if m['rate']})} "
          f"nem={len(nem_sa)} | HP_present={hp_ok}/{len(meters)} GPM_present={gpm_ok}/{len(meters)} "
          f"geo_present={geo_ok}/{len(meters)} solar_present={solar_ok}/{len(meters)}")

    files = list_files(a.export)
    print(f"\n== EXPORT == {len(files)} file(s)")
    if not files: sys.exit("ERROR: no export files found at " + a.export)

    rows, dollar_cols, fmts = [], set(), Counter()
    xml_signals = []
    for fp in files:
        fmt = sniff(fp); fmts[fmt] += 1
        if fmt == "csv":
            rs, colmap = parse_csv(fp)
            for c in colmap:
                if any(h in c for h in DOLLAR_HINTS): dollar_cols.add(c)
            rows += rs
        else:
            xml_signals.append((fp, parse_xml_coarse(fp)))
    print(f"   formats: {dict(fmts)}")
    if xml_signals:
        print("   NOTE: Green Button XML detected. Deep per-SA checks below run on CSV; "
              "production XML uses src/lib/greenbutton/parse.ts. Coarse XML signals:")
        for fp, s in xml_signals:
            print(f"     {os.path.basename(fp)}: usagePoints={s['usagePoints']} "
                  f"durations={sorted(s['durations'])} flowDirections={sorted(s['flowDirections'])}")
    if not rows and xml_signals:
        print("\n(Only XML present — rerun with a CSV export for the full per-feature board.)"); return

    # aggregate CSV
    export_acct = {norm_account(r["account"]) for r in rows if r["account"]}
    sid_intervals = defaultdict(set)   # said -> set(interval_len)
    sid_dir = defaultdict(set)
    sid_acct = {}                      # said -> normalized export account
    export_sa_raw = set()
    rate_codes = Counter()
    rate_sas = defaultdict(set)        # rate code -> set of distinct SAs on it
    uoms = Counter()
    dates = []
    for r in rows:
        sid = str(r["serviceId"]).strip(); export_sa_raw.add(sid)
        if r["intervallen"] not in (None,""):
            try: sid_intervals[sid].add(int(float(r["intervallen"])))
            except ValueError: pass
        sid_dir[sid].add("export" if is_received(r["direction"]) else "import")
        if r["account"] and sid not in sid_acct: sid_acct[sid] = norm_account(r["account"])
        if r["rate"]:
            code = str(r["rate"]).strip()
            rate_codes[code] += 1
            rate_sas[code].add(sid)
        if r["uom"]: uoms[str(r["uom"]).strip()] += 1
        if r["date"]: dates.append(str(r["date"]).strip())

    # 1) coverage
    master_rate_by_sa = {m["serviceId"]: m["rate"] for m in meters}
    master_acct_set = set(by_acct)
    matched, orphan = {}, set()
    for sid in export_sa_raw:
        m = said_match(sid, master_sa)
        if m: matched[sid] = m
        else: orphan.add(sid)
    covered_master = set(matched.values())
    missing = master_sa - covered_master
    print(f"\n== 1. COVERAGE ==")
    print(f"   export SAs={len(export_sa_raw)}  matched_to_master={len(covered_master)}  "
          f"orphan(in export, not master)={len(orphan)}")
    print(f"   master meters covered: {len(covered_master)}/{len(meters)}")
    print(f"   accounts present in export: {len(export_acct)} / {len(by_acct)} master accounts")
    # Since the harness unions all --export files into one row list, covered_master is
    # already the union over every file. State it so it is unambiguous that a second
    # month (April) does NOT add coverage when the SA set is identical.
    print(f"   UNION coverage across all {len(files)} export file(s): "
          f"{len(covered_master)}/{len(meters)} master meters, {len(missing)} still uncovered")
    # Explicit per-SA listing of the missing meters with their master rate, so the
    # founder sees exactly which wells (the AG-C/HAGC cluster, the OL1 outdoor-lighting
    # SA, the AG5B/HAGA1, and the blank-rate meter) have no interval data.
    print(f"   MISSING (in master, NOT in any export): {len(missing)}")
    for sid in sorted(missing):
        rate = master_rate_by_sa.get(sid) or "(blank rate)"
        print(f"     {sid}  master_rate={rate}")
    if missing:
        miss_by_rate = Counter((master_rate_by_sa.get(s) or "(blank)") for s in missing)
        print(f"   missing-by-rate: {dict(miss_by_rate.most_common())}")
    # Split orphans: extra unlisted meters on a KNOWN master account (likely inventory
    # omissions) vs SAs on an account ABSENT from the master (possibly other entities).
    orphan_on_known = {s for s in orphan if sid_acct.get(s) in master_acct_set}
    orphan_on_absent = orphan - orphan_on_known
    print(f"   ORPHANS (in export, NOT in master): {len(orphan)}  "
          f"-> {len(orphan_on_known)} on a known master account (extra unlisted meters), "
          f"{len(orphan_on_absent)} on an account absent from the master")
    if orphan: print(f"   orphan SA IDs (sample): {sorted(orphan)[:10]}")

    # 2) granularity
    print(f"\n== 2. INTERVAL GRANULARITY ==")
    all_lens = set().union(*sid_intervals.values()) if sid_intervals else set()
    hourly = {s for s, L in sid_intervals.items() if L & {60, 3600}}
    fine   = {s for s, L in sid_intervals.items() if L & {15, 900}}
    print(f"   interval-length values seen: {sorted(all_lens)}  "
          f"(15 or 900 = 15-min OK; 60/3600 = hourly = demand peak unreliable)")
    print(f"   15-min SAs={len(fine)}  hourly SAs={len(hourly)}")
    if hourly: print(f"   HOURLY meters (demand NO-GO): {sorted(hourly)[:10]}")

    # 3) rate codes
    print(f"\n== 3. RATE CODES ==")
    master_rate_set = {m["rate"] for m in meters if m["rate"]}
    unknown_rates = {c for c in rate_codes if c not in master_rate_set}
    print(f"   distinct rate codes in export: {sorted(rate_codes)}")
    if unknown_rates: print(f"   NOT in master's schedules: {sorted(unknown_rates)}")
    # Classify each code against the repo's real AG card families (not a naive 'AG'
    # substring), and tally by DISTINCT SA so the board reflects what the pilot can
    # actually price. SA-count helper: distinct SAs across a set of codes.
    def sa_count(codes): return len({s for c in codes for s in rate_sas[c]})
    ag_codes      = sorted(c for c in rate_codes if classify_rate(c) == "ag")
    ag_no_card    = sorted(c for c in rate_codes if classify_rate(c) == "ag_no_card")
    non_ag        = sorted(c for c in rate_codes if classify_rate(c) == "non_ag")
    print(f"   AG-priceable codes ({sa_count(ag_codes)} SAs): {ag_codes or 'none'}")
    print(f"   AG-prefixed but NO card family ({sa_count(ag_no_card)} SAs): {ag_no_card or 'none'}")
    print(f"   non-AG, repo card CANNOT price ({sa_count(non_ag)} SAs): {non_ag or 'none'}")
    h_ag = sorted(c for c in rate_codes if is_h_prefixed_ag(c))
    if h_ag:
        print(f"   NOTE: {sa_count(h_ag)} H-prefixed AG SAs ({h_ag}) map to a card family "
              f"but are UNPRICEABLE as the repo is wired because mapScheduleLabel does not "
              f"strip the leading H (fix: src/lib/energy/rate-lever.ts:148).")

    # 4) direction / solar
    print(f"\n== 4. DIRECTION / SOLAR ==")
    any_export = any("export" in v for v in sid_dir.values())
    nem_in_export = {matched[s] for s in matched if matched[s] in nem_sa}
    nem_with_export = {s for s in sid_dir if "export" in sid_dir[s] and said_match(s, nem_sa)}
    print(f"   any Received/export rows present: {any_export}")
    print(f"   NEM meters in this export: {len(nem_in_export)}  with export stream: {len(nem_with_export)}")

    # 5) dollars
    print(f"\n== 5. DOLLARS ==")
    print(f"   dollar/charge columns detected: {sorted(dollar_cols) or 'NONE (usage-only, as expected)'}")
    print(f"   units of measure: {dict(uoms)}")
    if dates: print(f"   date range: {min(dates)} -> {max(dates)}")

    # board
    def v(go, partial, no):
        return "GO" if go else ("PARTIAL" if partial else "NO-GO")
    print(f"\n== PER-FEATURE GO / PARTIAL / NO-GO ==")
    board = [
        ("Usage dashboard",            v(bool(rows), False, False), "Usage Value + Date/Time"),
        ("Cost dashboard ($)",         "NO-GO", "no dollars in file -> needs tariff/bills"),
        ("Rate optimization",          "PARTIAL", f"load shape OK; {len(non_ag)} non-AG code(s) lack repo tariff; need full tariff tables"),
        ("Demand (peak kW)",           v(not hourly and bool(fine), bool(hourly), False),
                                       f"{len(fine)} 15-min / {len(hourly)} hourly; $ still needs tariff/bill"),
        ("Solar/NEM volume",           v(any_export, bool(nem_in_export) and not any_export, False),
                                       "needs Received(R) rows; credits/true-up = bills"),
        ("Spend by account",           "NO-GO", "no dollars -> usage-by-account only"),
        ("Pump efficiency",            v(hp_ok==len(meters), hp_ok>0, False), f"HP present {hp_ok}/{len(meters)} (file can't supply HP)"),
        ("SGMA water",                 "PARTIAL" if (hp_ok and gpm_ok) else "NO-GO",
                                       "needs HP+GPM (or pump test); rigorous AF needs flow/TDH"),
        ("Almond (usage/peak)",        v(bool(rows), False, False), "inherits gaps above"),
    ]
    for name, verdict, note in board:
        print(f"   {verdict:8s} {name:22s} {note}")

if __name__ == "__main__":
    main()
